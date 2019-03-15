from cpppo.server.enip.get_attribute import proxy
import smtplib
from threading import Timer
import traceback
import operator
from twilio.rest import Client
from openpyxl import load_workbook
from datetime import datetime
from time import time
from time import ctime
from time import sleep
import configparser

# define the operators that are available. These text operators are configured in the alarm_configuration.xlsx file.
ops = {">": operator.gt, "<": operator.lt, "=": operator.eq, "!=": operator.ne}  # etc.
str_to_bool = {'0': False, '1': True}


#This PLC class is used to maintain a CIP connection to the PLC's and makes it easy to read tag values from those PLC's.
class PLC:
    def __init__(self, ip, slot):
        self.ip = ip
        self.slot = slot
        if str(self.slot) != '0':
            self.proxy = proxy(self.ip, route_path=[{'link': 1, 'port': int(self.slot)}], timeout=3)
        else:
            self.proxy = proxy(self.ip, timeout=3)

    def read_value(self, tag):
        log("Reading - IP:" + str(self.ip) + " Slot:" + str(self.slot) + " Tag:" + tag)

        try:
            i = 0
            while True:
                if i > 0:
                    sleep(.01)
                i += 1
                value = next(self.proxy.read(tag, checking=True))
                if value:
                    break
                if i > 3:
                    value = [0]
                    log("Value failed to read after " + str(i) + " attempts - IP:" + str(self.ip) + " Slot:" + str(self.slot) + " Tag:" + tag)
                    break
            return value[0]
        except:
            self.proxy.close_gateway()
            log("There was an error reading the tag: " + tag + "\r\nip:" + self.ip + "\r\nSlot: " + str(self.slot) + "\r\n" + traceback.format_exc())
            send_email(config['DEFAULT']['debug_subject'], "IP: " + self.ip + "\r\nTag:" + tag + "\r\nSlot:" + str(self.slot) + '\r\n' + traceback.format_exc(), [config['DEFAULT']['debug_email']])


class ALARM:
    def __init__(self, ip, slot, tag, operator, value, message, process_value, email, sms):
        self.ip = ip
        self.slot = slot
        self.tag = tag
        # string defining the operator to be used for the alarm condition
        self.operator = operator
        # function defined by the 'operator' column looking up in the 'ops' dictionary of functions.
        self.ops = ops[operator]
        self.value = value
        self.message = message
        self.process_value = process_value
        self.email = email
        self.sms = sms
        # variable/more complicated tag storage items
        self.last_notification_time = time()
        # list of the last times the alarm was triggered
        self.history = list()
        self.alarm_count = 0
        self.email_alarm_subject = "Alert! : " + self.tag + " " + self.operator + " " + str(self.value)
        self.email_recovered_subject = "Recovered! : " + self.tag + " " + self.operator + " " + str(self.value)
        # reads the current value of the alarm tag
        tag_value = connections[self.ip][self.slot].read_value(self.tag)
        # initial value, so it doesn't send email alerts when the script is first started.
        self.triggered = self.ops(tag_value, self.value)

    def check_alarm(self):
        try:
            print("Checking alarm: " + self.tag)
            # read the value from the plc.
            tag_value = connections[self.ip][self.slot].read_value(self.tag)
            # check to make sure the value read successfully and if the alarm condition is met.
            in_alarm = self.ops(tag_value, self.value) and (tag_value is not None)
            # if the alarm has already been triggered
            if in_alarm and not self.triggered:
                if time() > self.last_notification_time + int(config['DEFAULT']['max_notification_frequency']):
                    self.history.append("Alarmed at:  \t" + ctime(int(time())))
                    self.history = self.history[-20:]
                    current_process_value = str(connections[self.ip][self.slot].read_value(self.process_value))
                    message = self.message + '\r\n\r\n' + "Process variable: " + str(self.process_value) + " = " + str(current_process_value) + "\r\nAlarm History:\r\n" + '\r\n'.join(self.history)
                    email_sent = send_email(self.email_alarm_subject, message, self.email)
                    sms_sent = send_sms(self.sms, self.email_alarm_subject + ":" + self.message, config['DEFAULT']['twilio_id'], config['DEFAULT']['twilio_pw'], config['DEFAULT']['twilio_number'])
                    if email_sent or sms_sent:
                        self.last_notification_time = time()
                        self.triggered = True

            elif not in_alarm and self.triggered:
                self.history.append("Recovered at:\t"+ctime(int(time())))
                self.history = self.history[-20:]
                current_process_value = str(connections[self.ip][self.slot].read_value(self.process_value))
                message = self.message + '\r\n\r\n' + "Process variable: " + self.process_value + " = " + str(current_process_value) + "\r\nAlarm History:\r\n" + '\r\n'.join(self.history)
                email_sent = send_email(self.email_recovered_subject, message, self.email)
                sms_sent = send_sms(self.sms, self.email_recovered_subject + ":" + self.message, config['DEFAULT']['twilio_id'], config['DEFAULT']['twilio_pw'])
                if email_sent or sms_sent:
                    self.triggered = False

        except:
            log("There was an error checking the alarm:" + traceback.format_exc())
            send_email(config['DEFAULT']['debug_subject'], traceback.format_exc(), [config['DEFAULT']['debug_email']])


# Write to a log file
def log(text_to_log):
    with open('log.txt', 'a') as f:
        print(str(datetime.now()) + "\r\n" + str(text_to_log) + '\r\n\r\n')
        f.write(str(datetime.now()) + "\r\n" + str(text_to_log) + '\r\n\r\n')


# Use twilio to send an sms message
def send_sms(deliver_to, content, twilio_id, twilio_pw, twilio_number):
    deliver_to = list(set(deliver_to))
    log(deliver_to)
    log(content)
    try:
        if deliver_to != "" and deliver_to is not None:
            client = Client(twilio_id, twilio_pw)
            for phone_number in deliver_to:
                log(phone_number)
                client.messages.create(to=phone_number, from_=twilio_number, body=content)
    except:
        var = traceback.format_exc()
        send_email(config['DEFAULT']['debug_subject'], "Deliver to: " + deliver_to + "\r\nContent: " + content + "\r\n" + var, config['DEFAULT']['debug_email'])
        log("Deliver to: " + deliver_to + "\r\nContent: " + content + "\r\n" + var)
        return True
    return True


# send an email using the settings in the config file.
def send_email(subject, body, email_to):
    email_to = list(set(email_to)) #This makes sure that emails don't exist twice in the config, otherwise duplicate emails could be sent.
    message = "From: "+config['DEFAULT']['email_from']+"\nTo: " + ", ".join(email_to)+"\nSubject: "+subject+"\n\n"+body
    log(message)
    try:
        server = smtplib.SMTP(config['DEFAULT']['email_server'], int(config['DEFAULT']['email_port']))
        server.ehlo()
        server.starttls()
        if config['DEFAULT']['email_user'] != "" and config['DEFAULT']['email_password'] != "":
            server.login(config['DEFAULT']['email_user'], config['DEFAULT']['email_password'])
        server.sendmail(config['DEFAULT']['email_from'], email_to, message.encode('ascii','ignore').decode('ascii','ignore'))
        server.close()
        return True
    except:
        var = traceback.format_exc()
        log("Subject: " + subject + "\r\nBody: " + body + "\r\n" + var)
        return False


# loads the alarm configuration and the notification subscription settings.
# Auto-reloads so the application doesn't have to restart.
def load_alarm_definitions():
    global alarms
    global connections
    try:
        wb = load_workbook('alarm_configuration.xlsx')
        ws = wb['alarm_configuration']
        for row in ws['A4:H{}'.format(ws.max_row)]:
            if row[0].value in alarms:
                alarm = alarms[row[0].value]
                # check if there are differences between the spreadsheet and what has been configured for the alarm.
                if not (alarm.ip == row[1].value
                        and alarm.slot == row[2].value
                        and alarm.tag == row[3].value
                        and alarm.operator == row[4].value
                        and alarm.value == row[5].value
                        and alarm.message == row[6].value
                        and alarm.process_value == row[7].value):
                    alarm.ip = row[1].value
                    alarm.slot = row[2].value
                    alarm.tag = row[3].value
                    alarm.operator = row[4].value
                    alarm.value = row[5].value
                    alarm.message = row[6].value
                    alarm.process_value = row[7].value
                    alarm.triggered = '0'
                    alarm.email = list()
                    alarm.sms = list()
            else:
                if row[1].value not in connections:
                    connections[row[1].value] = dict()
                if row[2].value not in connections[row[1].value]:
                    connections[row[1].value][row[2].value] = PLC(row[1].value, row[2].value)
                alarm = ALARM(row[1].value,  # ip address
                              row[2].value,  # slot
                              row[3].value,  # tag name
                              row[4].value,  # operator
                              row[5].value,  # value
                              row[6].value,  # message
                              row[7].value,  # process value
                              list(),  # list of emails, to be populated below
                              list())  # list of sms numbers, to be populated below
            alarms[row[0].value] = alarm
        # iterate through columns starting with 9 which is "J"
        for user in ws.iter_cols(min_row=1, min_col=10, max_col=ws.max_column):
            # iterates through the rows, 1 for each alarm. Adds the users details if the value is marked.
            for r in range(3, ws.max_row):
                alarm = alarms[ws['A:A'][r].value]
                if user[r].value == 1 or user[r].value == 3:
                    alarm.email.append(user[1].value)
                if user[r].value == 2 or user[r].value == 3:
                    alarm.sms.append(user[0].value)
                alarms[ws['A:A'][r].value] = alarm
    except:
        log("There was an error loading the alarm configuration" + traceback.format_exc())
        send_email(config['DEFAULT']['debug_subject'], traceback.format_exc(), [config['DEFAULT']['debug_email']])
    # Re-loads the alarm definitions every 60 seconds.
    Timer(60, load_alarm_definitions).start()


# check all the alarms, start another alarm check to run at the defined rate in the future.
def check_alarms():
    start_time = time()
    try:
        for index in alarms.keys():
            alarms[index].check_alarm()
    except:
        log("There was an error checking the alarms:" + traceback.format_exc())
        send_email(config['DEFAULT']['debug_subject'], traceback.format_exc(), [config['DEFAULT']['debug_email']])
    delay = int(int(config['DEFAULT']['frequency']) - (time() - start_time))
    if delay < 1:
        delay = 1
    Timer(delay, check_alarms).start()


if __name__ == '__main__':
    connections = dict()
    alarms = dict()
    # load the global configuration from config.txt
    config = {}
    # reads the configuration from config.txt stored in the same location.
    config = configparser.ConfigParser()
    config.read('config.ini')
    
    send_email('CIP Tag Monitor Started', str(datetime.now()), [config['DEFAULT']['debug_email']])
    # start the alarm checking function
    load_alarm_definitions()
    check_alarms()